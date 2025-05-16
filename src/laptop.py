"""
Copyright (c) 2023 The uos_sess6072_build Authors.
Authors: Miquel Massot, Blair Thornton, Sam Fenton
All rights reserved.
Licensed under the BSD 3-Clause License.
See LICENSE.md file in the project root for full license information.
"""
from Libraries import *
from Tools import *
import numpy as np
import argparse
import time
import openpyxl
import pickle
import sys

from datetime import datetime
from drivers.aruco_udp_driver import ArUcoUDPDriver
from zeroros import Subscriber, Publisher
from zeroros.messages import LaserScan, Vector3Stamped, Pose, PoseStamped, Header, Quaternion
from zeroros.datalogger import DataLogger
from zeroros.rate import Rate

from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.gaussian_process.kernels import ConstantKernel, RBF

from matplotlib import pyplot as plt
from openpyxl import load_workbook


# add more libraries here
N = 0
E = 1
G = 2
DOTX = 3
DOTG = 4
class LaptopPilot:
    def __init__(self, simulation):
        # network for sensed pose
        aruco_params = {
            "port": 50000,  # Port to listen to (DO NOT CHANGE)
            "marker_id": 21,  # Marker ID to listen to (CHANGE THIS to your marker ID)            
        }
        self.robot_ip = "192.168.90.1"
        self.plotGroundtruth = None
        # handles different time reference, network amd aruco parameters for simulator
        self.sim_time_offset = 0 #used to deal with webots timestamps
        self.sim_init = False #used to deal with webots timestamps
        self.simulation = simulation
        self.aruco = False

        if self.simulation:
            self.robot_ip = "127.0.0.1"          
            aruco_params['marker_id'] = 0  #Ovewrites Aruco marker ID to 0 (needed for simulation)
            self.sim_init = True #used to deal with webots timestamps
            self.plotGroundtruth = True
        
        print("Connecting to robot with IP", self.robot_ip)
        self.aruco_driver = ArUcoUDPDriver(aruco_params, parent=self)
        self.initialise_pose = True # False once the pose is initialised

        ############# INITIALISE ATTRIBUTES ##########        
        #>Modelling<#
        ################
        # path
        self.path_velocity = 0.04
        self.path_acceleration = 0.1/3
        self.path_radius = 0.3
        self.accept_radius = 0.2
        lapx = [0,1.4,1.4,0]
        lapy = [0,0,1.4,1.4]
        # lapx = [0,-0.1]
        # lapy = [0,-0.1]
        self.northings_path = lapx+lapx+lapx
        self.eastings_path = lapy+lapy+lapy
        self.relative_path = True #False if you want it to be absolute  
        # modelling parameters
        wheel_distance = 0.174/2 # m 
        wheel_diameter = 0.07 # m
        self.ddrive = ActuatorConfiguration(wheel_distance, wheel_diameter) #look at your tutorial and see how to use this
    
        # self.lidar_rangenoise = 0.000025
        # self.lidar_anglenoise = 0.0003
        self.lidar_rangenoise = 0 
        self.lidar_anglenoise = 0
        self.new_lidar = None

        # control parameters        
        self.tau_s = 0.5 # s to remove along track error
        self.L = 0.2 # m distance to remove normal and angular error
        self.v_max = 0.6 # m/s fastest the robot can go
        self.w_max = np.deg2rad(120) # fastest the robot can turn
        self.timeout = 10 #s
        
        self.initialise_control = True # False once control gains is initialised 

        # graphslam
        self.stop_to_think = False
        self.za_warudo_part_one = False
        self.za_warudo_part_two = False
        self.landmark_acceptance_radius = 1
        self.count = 0
        self.p_ = 0
        self.sigma_ = 0
        self.pose_groundtruth = []
        
        # classifiers
        self.cornerClassifier = None
        self.landmark = None # For sending to show_laptop
        self.new_landmark = None

        # model pose
        self.est_pose_northings_m = None
        self.est_pose_eastings_m = None
        self.est_pose_yaw_rad = None

        #motion model variables
        self.state = None
        self.sensor_measurement = None

        # kalman filter
        self.jacobian = None
        self.covariance = None
        self.uncertainty = None
        self.dptracker = []
        #>Communication>#
        #################
    
        # measured pose
        self.measured_pose_timestamp_s = None
        self.measured_pose_northings_m = None
        self.measured_pose_eastings_m = None
        self.measured_pose_yaw_rad = None

        # wheel speed commands
        self.cmd_wheelrate_right = None
        self.cmd_wheelrate_left = None 

        # encoder/actual wheel speeds
        self.measured_wheelrate_right = None
        self.measured_wheelrate_right_next = None
        self.measured_wheelrate_left = None

        # measured ground speeds
        self.groundtruth_northings = None
        self.groundtruth_eastings = None
        self.groundtruth_yaw = None   

        # lidar
        self.lidar_timestamp_s = None
        self.lidar_data = None
        lidar_xb = 0 # location of lidar centre in b-frame primary axis ########################(changed)
        lidar_yb = 0.1 # location of lidar centre in b-frame secondary axis ###################(Changed)
        self.lidar = RangeAngleKinematics(lidar_xb,lidar_yb) ####################(changed)

        # Excel export
        self.export_data = None
        self.ref_pose_worksheet = LaptopPilot.createExcelFile("output_data.xlsx")

        # Create variable for plotting ground truth and reference position
        self.p_reference_tracker = None
        self.p_groundtruth_tracker = None
        self.p_gt_path = []

        
        ###############################################################        

        self.datalog = DataLogger(log_dir="logs")

        # Wheels speeds in rad/s are encoded as a Vector3 with timestamp, 
        # with x for the right wheel and y for the left wheel.        
        self.wheel_speed_pub = Publisher(
            "/wheel_speeds_cmd", Vector3Stamped, ip=self.robot_ip
        )

        self.true_wheel_speed_sub = Subscriber(
            "/true_wheel_speeds",Vector3Stamped, self.true_wheel_speeds_callback,ip=self.robot_ip,
        )
        self.lidar_sub = Subscriber(
            "/lidar", LaserScan, self.lidar_callback, ip=self.robot_ip
        )
        self.groundtruth_sub = Subscriber(
            "/groundtruth", Pose, self.groundtruth_callback, ip=self.robot_ip
        )
    
    class createExcelFile:
        def __init__(self, filename = "reference.xlsx"):
            # Creates workbook and worksheet to modify
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.filename = filename
            self.workbook.save(filename)

            # Initialise variable to store data
            self.dataLine = []

        def extend_data(self, data):
            # adds to list self.dataLine
            data = change_to_list(data)
            self.dataLine.extend(data)

        def export_to_excel(self):
            # appends dataLine to sheet and saves file
            self.workbook = load_workbook(self.filename)
            self.worksheet.append(self.dataLine)       
            self.workbook.save(self.filename)
            self.workbook.close
            self.dataLine = []

    def true_wheel_speeds_callback(self, msg):
        # print("Received sensed wheel speeds: R=", msg.vector.x,", L=", msg.vector.y)
        # update wheel rates
        self.measured_wheelrate_right = self.measured_wheelrate_right_next
        self.measured_wheelrate_right_next = msg.vector.x
        self.measured_wheelrate_left = msg.vector.y

        self.datalog.log(msg, topic_name="/true_wheel_speeds")

    def lidar_callback(self, msg):
        # This is a callback function that is called whenever a message is received        
        print("Received lidar message", msg.header.seq)        
        if self.sim_init == True:
            self.sim_time_offset = datetime.utcnow().timestamp()-msg.header.stamp
            self.sim_init = False     

        msg.header.stamp += self.sim_time_offset
        ###############(imported)#########################
        self.lidar_timestamp_s = msg.header.stamp #we want the lidar measurement timestamp here
        self.lidar_data = np.zeros((len(msg.ranges), 2)) #specify length of the lidar data
        self.lidar_data[:,0] = msg.ranges # use ranges as a placeholder, workout northings in Task 4
        self.lidar_data[:,1] = msg.angles # use angles as a placeholder, workout eastings in Task 4

        self.raw_lidar = self.lidar_data[:120,:]
        ###############(imported)#########################
        self.datalog.log(msg, topic_name="/lidar")

        ###############(imported)#########################
        # b to e frame
        p_eb = Vector(3)
        p_eb[0] = self.est_pose_northings_m #robot pose northings (see Task 3)
        p_eb[1] = self.est_pose_eastings_m #robot pose eastings (see Task 3)
        p_eb[2] = self.est_pose_yaw_rad #robot pose yaw (see Task 3)

        # m to e frame
        self.lidar_data = np.zeros((len(msg.ranges), 2))        
                    
        z_lm = Vector(2)        
        # for each map measurement
        for i in range(len(msg.ranges)):
            z_lm[0] = msg.ranges[i]
            z_lm[1] = msg.angles[i]
                
            t_em = self.lidar.rangeangle_to_loc(p_eb, z_lm) # see tutotial

            self.lidar_data[i,0] = t_em[0]
            self.lidar_data[i,1] = t_em[1]
        
        self.new_lidar = True



    def groundtruth_callback(self, msg):
        """This callback receives the odometry ground truth from the simulator."""
        self.groundtruth_northings = msg.position.x
        self.groundtruth_eastings = msg.position.y 
        _, _, self.groundtruth_yaw = msg.orientation.to_euler()  
        self.datalog.log(msg, topic_name="/groundtruth")

    def groundtruth_update(self):
            p_gt = Vector(3)
            p_gt[0] = self.groundtruth_northings
            p_gt[1] = self.groundtruth_eastings
            p_gt[2] = self.groundtruth_yaw

            return p_gt
    
    def pose_parse(self, msg, aruco = False):
        # parser converts pose data to a standard format for logging
        time_stamp = msg[0]

        if aruco == True:
            if self.sim_init == True:
                self.sim_time_offset = datetime.utcnow().timestamp()-msg[0]
                self.sim_init = False                                         
                
            # self.sim_time_offset is 0 if not a simulation. Deals with webots dealing in elapse timeself.sim_time_offset
            print(
                "Received update from",
                datetime.utcnow().timestamp() - msg[0] - self.sim_time_offset,
                "seconds ago",
            )
            time_stamp = msg[0] + self.sim_time_offset                

        pose_msg = PoseStamped() 
        pose_msg.header = Header()
        pose_msg.header.stamp = time_stamp
        pose_msg.pose.position.x = msg[1]
        pose_msg.pose.position.y = msg[2]
        pose_msg.pose.position.z = 0

        quat = Quaternion()        
        if self.simulation == False and aruco == True: quat.from_euler(0, 0, np.deg2rad(msg[6]))
        else: quat.from_euler(0, 0, msg[6])
        pose_msg.pose.orientation = quat        
        return pose_msg

    def generate_trajectory(self):
    # pick waypoints as current pose relative or absolute northings and eastings
        if self.relative_path == True:
            for i in range(len(self.northings_path)):
                self.northings_path[i] += self.est_pose_northings_m #offset by current northings
                self.eastings_path[i] += self.est_pose_eastings_m #offset by current eastings

        # convert path to matrix and create a trajectory class instance
        C = l2m([self.northings_path, self.eastings_path])        
        self.path = TrajectoryGenerate(C[:,0],C[:,1])        
            
        # set trajectory variables (velocity, acceleration and turning arc radius)
        self.path.path_to_trajectory(self.path_velocity, self.path_acceleration) #velocity and acceleration
        self.path.turning_arcs(self.path_radius) #turning radius
        self.path.wp_id=0 #initialises the next waypoint
        ####################  (^^^^^^^imported^^^^^^)

    def run(self, time_to_run=-1):
        self.start_time = datetime.utcnow().timestamp()
        
        try:
            r = Rate(10.0)
            while True:
                current_time = datetime.utcnow().timestamp()
                if time_to_run > 0 and current_time - self.start_time > time_to_run:
                    print("Time is up, stopping…")
                    break
                self.infinite_loop()
                r.sleep()
        except KeyboardInterrupt:
            print("KeyboardInterrupt received, stopping…")
        except Exception as e:
            print("Exception: ", e)
        finally:
            self.lidar_sub.stop()
            self.groundtruth_sub.stop()
            self.true_wheel_speed_sub.stop()
    
    def initialise_robot(self):
        # Create initial state, covariance and position estimate
        self.state = Vector(3)
        if self.aruco == True:
            self.state[0] = self.measured_pose_northings_m 
            self.state[1] = self.measured_pose_eastings_m  
            self.state[2] = self.measured_pose_yaw_rad
        elif self.simulation == True:
            self.state[0] = 0.3
            self.state[1] = 0.3
            self.state[2] = 0
        else:
            self.state[0] = 0.0
            self.state[1] = 0.0
            self.state[2] = 0.0
        self.update_estimated_pose()


    def position_sensor_update(self):
        # Sample position data from Aruco
        sensor_measurement = Vector(3)
        sensor_measurement[N] = self.measured_pose_northings_m
        sensor_measurement[E] = self.measured_pose_eastings_m
        sensor_measurement[G] = self.measured_pose_yaw_rad

        return sensor_measurement

    def lidar_addnoise(self, lidardata):
        rangenoise = add_noise(self.lidar_rangenoise,0,len(lidardata[:,0]))
        anglenoise = add_noise(self.lidar_anglenoise,0,len(lidardata[:,1]))

        lidardata[:,0] += rangenoise
        lidardata[:,1] += anglenoise

        return lidardata

        
    class uncertaintyMatrices:  
        #Class to keep track of model uncertainty
        def get_initial_uncertainty(self):
            # Create process uncertainty matrix
            # sigma = Matrix(3,3) 
            # sigma[0,0]=0.01**2
            # sigma[0,1]=0.001**2
            # sigma[1,0]=0.001**2
            # sigma[1,1]=0.001**2
            # sigma[0,2]=0.01**2
            # sigma[1,2]=0.001**2
            # sigma[2,0]=0.001**2
            # sigma[2,1]=0.001**2
            # sigma[2,2]=0.01**2

            sigma = Matrix(3,3) 
            sigma[0,0]=0.005
            sigma[0,1]=0.0005
            sigma[1,0]=0.0005
            sigma[1,1]=0.005
            sigma[0,2]=0.0005
            sigma[1,2]=0.0005
            sigma[2,0]=0.0005
            sigma[2,1]=0.0005
            sigma[2,2]=0.005

            return sigma

        def get_process_uncertainty3x3(self):
            #Motion model linear noise due to v and w
            sigma_motion=Matrix(3,2)
            sigma_motion[0,0]= 0.0155**2 # impact of v linear velocity on x           #Task
            sigma_motion[0,1]= np.deg2rad(0.012)**2# impact of w angular velocity on x
            sigma_motion[1,0]= 0.0155**2# impact of v linear ve   locity on y
            sigma_motion[1,1]=np.deg2rad(0.012)**2 # impact of w angular velocity on y
            sigma_motion[2,0]= 0.00155**2 # impact of v linear velocity on gamma
            sigma_motion[2,1]=np.deg2rad(0.0012)**2 # impact of w linear velocity on gamma
            
            return sigma_motion

        def get_p_sensor_uncertainty(self):
            # Create position sensor uncertainty matrix
            Q = Identity(5)

            Q[N, N] = 0.00**2
            Q[E, E] = 0.00**2

            return Q
        
        def get_yaw_sensor_uncertainty(self):
            # Create yaw sensor uncertainty matrix
            Q = Identity(5)

            Q[G, G] = np.deg2rad(0.0)**2

            return Q
        
        def get_lidar_uncertainty(self):
            # Get lidar uncertainty
            Ql = Identity(2)

            Ql[N,N] = 0.1**2 #range
            Ql[N,E] = 0
            Ql[E,N] = np.deg2rad(5)**2 #Angle from range
            Ql[E,E] = 0
            # Ql[N,N] = 0
            # Ql[N,E] = 0#
            # Ql[E,N] = 0
            # Ql[E,E] = 0

            return Ql

    def update_estimated_pose(self):
        # Update estimate variable for logging
        self.est_pose_northings_m = self.state[0,0]
        self.est_pose_eastings_m = self.state[1,0]
        self.est_pose_yaw_rad = self.state[2,0]
    
    def initialise(self):
        if self.initialise_pose == True:
            time.sleep(0.5)
            # set initial measurements
            self.initialise_robot()
            self.uncertainty = LaptopPilot.uncertaintyMatrices()
            self.covariance = self.uncertainty.get_initial_uncertainty()

            self.generate_trajectory()

            #Graphslam
            self.cornerClassifier = Classifier()
            self.cornerClassifier.train_classifier('corner')

            self.graph = graphslam_frontend()
            self.graph.anchor(self.covariance)

            H_eb = HomogeneousTransformation(self.state[0:2], self.state[2])
            self.reset_previous_state(H_eb)
            self.new_cov = Identity(3)

            # get current time and determine timestep
            self.t_prev = datetime.utcnow().timestamp() #initialise the time
            self.t = 0 #elapsed time
            time.sleep(0.1) #wait for approx a timestep before proceeding
            
            # path and tragectory are initialised
            self.initialise_pose = False
    
    def robo_stop(self):
        wheel_speed_msg = Vector3Stamped()
        wheel_speed_msg.vector.x = 0 # Right wheelspeed rad/s
        wheel_speed_msg.vector.y = 0 # Left wheelspeed rad/s

        self.cmd_wheelrate_right = wheel_speed_msg.vector.x
        self.cmd_wheelrate_left = wheel_speed_msg.vector.y

        self.wheel_speed_pub.publish(wheel_speed_msg)
        self.datalog.log(wheel_speed_msg, topic_name="/wheel_speeds_cmd")

    def reset_previous_state(self, H_eb):
        self.p_ = self.state
        self.sigma_ = self.covariance
        self.dp_ = 0
        self.sigma_ = H_eb.H_R @ self.sigma_ @ H_eb.H_R.T

    def save_object(self, object, filename):
        with open(filename, "wb") as f:
            pickle.dump(object, f)

    def classify_lidar(self):
        if self.new_lidar == True:
                self.new_lidar = False
                # self.raw_lidar = self.lidar_addnoise(self.raw_lidar)
                observation = GPC_input_output(self.raw_lidar, None)
                #Check for corners
                corner_probability = self.cornerClassifier.classifier.predict_proba([observation.data_filled[:, 0]])
                # Remove lidar points with a range below 5cm (for real life lidar)
                observation.data_filled = observation.data_filled[observation.data[:, 0] > 0.01]
                print(corner_probability)  
                # Only check for corners if there are at least 10 points
                if corner_probability[0][0] > 0.5 and len(observation.data_filled) > 10:  
                    label = (self.cornerClassifier.classifier.classes_[np.argmax(corner_probability)])
                    print(label)
                    # Check for corners                
                    z_lm = Vector(2)
                    z_lm[0], z_lm[1], loc = find_corner(observation, 0.003)
                    z_lm_checker = z_lm.shape == (2,1) and np.issubdtype(z_lm.dtype, np.floating) and not np.isnan(z_lm).any()
                    # If the corner exists and its coordinates are not nan
                    if loc is not None and z_lm_checker:
                        # Record landmarks for graphslam
                        self.landmark = self.lidar.rangeangle_to_loc(self.state, z_lm)
                        if not np.isnan(self.landmark).any():
                            print(label, "at", self.landmark)
                            self.new_landmark = True
                else:
                    self.landmark = None

    def infinite_loop(self):
        """Main control loop

        Your code should go here.
        """
        # > Sense < #
        # get the latest position measurements
        if self.aruco:
            aruco_pose = self.aruco_driver.read()    

            if aruco_pose is not None:
                # converts aruco date to zeroros PoseStamped format
                msg = self.pose_parse(aruco_pose, aruco = True)
                # reads sensed pose for local use
                self.measured_pose_timestamp_s = msg.header.stamp
                self.measured_pose_northings_m = msg.pose.position.x
                self.measured_pose_eastings_m = msg.pose.position.y
                _, _, self.measured_pose_yaw_rad = msg.pose.orientation.to_euler()        
                self.measured_pose_yaw_rad = self.measured_pose_yaw_rad % (np.pi*2) # manage angle wrapping

                # logs the data            
                self.datalog.log(msg, topic_name="/aruco")
                self.initialise()
        elif self.measured_wheelrate_right is not None:
            self.initialise()
        # initialisation step
    
        if self.initialise_pose != True:  
            
             # > Receive < #
            #################################################################################
            # convert true wheel speeds in to twist
            q = Vector(2)            
            q[0] = self.measured_wheelrate_right # wheel rate rad/s (measured)
            q[1] = self.measured_wheelrate_left # wheel rate rad/s (measured)
            u = self.ddrive.fwd_kinematics(q)    
            
            #determine the time step

            t_now = datetime.utcnow().timestamp()        

            if self.za_warudo_part_one is True:
                self.za_warudo_part_one = False
                self.za_warudo_part_two = True
                self.t_prev = t_now
                dt = 0.001
            else:
                dt = t_now - self.t_prev #timestep from last estimate
                self.t += dt #add to the elapsed time


             # > Think < #
            ################################################################################
            #gwound twuth pwotter
            if self.simulation:
                p_gt = self.groundtruth_update()
            elif self.aruco:
                p_gt = self.position_sensor_update()
            else:
                p_gt = self.state
            
            ######################################
            # Lidar observation 
            ######################################
            #If a new observation is available
            
            self.classify_lidar()
            #####################################
            # Graphslam frontend
            #####################################
            # Get current pose to pose uncertainty
            # Save previous pose and smegma
            H_eb = HomogeneousTransformation(self.state[0:2], self.state[2])
            sigma_motion = self.uncertainty.get_process_uncertainty3x3()
            sigma_lidar = self.uncertainty.get_lidar_uncertainty()
            
            # Progress motion model
            self.state, self.covariance, dp, _ =  rigid_body_kinematics(self.state,u,dt=dt,mu_gt=None,sigma_motion=sigma_motion,sigma_xy=self.covariance)
            self.dp_ += dp
            self.dptracker.append(dp)
            # Log landmark

            if self.new_landmark is True:
                self.new_landmark = False
                _, _, t_lm, sigma_observe_xy = self.lidar.loc_to_rangeangle(p_eb=self.state,t_em=self.landmark,sigma_observe=sigma_lidar)
                new_id, landmark_id = check_landmarks(self.landmark, self.graph, self.landmark_acceptance_radius)
                print(landmark_id)
                # Check distance to other landmarks
                self.graph.observation(self.landmark, sigma_observe_xy, landmark_id, t_lm)
                # Check if landmark is older than 1 observation
                if new_id is False and landmark_id != self.graph.landmark_id_array[-2]:
                    # Stop robot motion cause it needsa think
                    self.robo_stop()
                    self.stop_to_think = True
                    self.za_warudo_part_one = True
                else:
                    self.graph.motion(self.p_, self.sigma_, self.dp_, final=False)
                    self.pose_groundtruth.append(self.groundtruth_update())
                    self.reset_previous_state(H_eb)
                self.count = 0
            # If no landmarks and moving, business as usual 
            elif u.all() != 0 and self.count > 10:
                self.graph.motion(self.p_, self.sigma_, self.dp_, final=False)
                self.pose_groundtruth.append(self.groundtruth_update())
                self.reset_previous_state(H_eb)
                self.count = 0
            self.count += 1

            # Optimise graph
            if self.za_warudo_part_two:
                    self.za_warudo_part_two = False
                    self.robo_stop()
                    self.graph.motion(self.p_, self.sigma_, self.dp_, final=True)
                    self.reset_previous_state(H_eb)
                    self.pose_groundtruth.append(self.groundtruth_update())
                    self.graph.construct_graph() 
                    init_graph = self.graph
                    np.savetxt("output.csv", self.graph.H, delimiter=",", fmt='%d')
                    self.save_object(init_graph,filename="init_graph.pkl")
                    self.save_object(self.pose_groundtruth, "grdtruth.pkl")
                    # Optimise graph and update position
                    self.graph = juice_graph(self.graph)
                    print("old state", self.state)
                    self.state = self.graph.pose[-1:][0]
                    print("new state", self.state)
                    self.save_object(self.graph,filename="final_graph.pkl")
                    #

                    
            # update for show_laptop.py            
            self.update_estimated_pose()
            
            msg = self.pose_parse([datetime.utcnow().timestamp(),self.est_pose_northings_m,self.est_pose_eastings_m,0,0,0,self.est_pose_yaw_rad])
            self.datalog.log(msg, topic_name="/est_pose")      

            # feedforward control: check wp progress and sample reference trajectory
            self.path.wp_progress(self.t, self.state[0:3],self.accept_radius,2,self.timeout) # fill turning radius
            p_ref, u_ref = self.path.p_u_sample(self.t) #sample the path at the current elapsetime (i.e., seconds from start of motion modelling)
            self.p_reference_tracker = p_ref[0:2,0]

            if self.stop_to_think == False:
                self.t_prev = t_now #update the previous timestep for the next loop
                # > Control < #
                ################################################################################
                # feedback control: get pose change to desired trajectory from body
                dp = p_ref - self.state[0:3] #compute difference between reference and estimated pose in the $e$-frame

                dp[2] = (dp[2] + np.pi) % (2 * np.pi) - np.pi # handle angle wrapping for yaw

                ds = Inverse(H_eb.H_R) @ dp # rotate the $e$-frame difference to get it in the $b$-frame (Hint: dp_b = H_be.H_R @ dp_e)

                # compute control gains for the initial condition (where the robot is stationalry)
                self.k_s = 1/self.tau_s #ks
                if self.initialise_control == True:
                    self.k_n = 0 #kn
                    self.k_g = 0 #kg
                    self.initialise_control = False # maths changes a bit after the first iteration

                # update the controls
                du = feedback_control(ds, self.k_s, self.k_n, self.k_g)

                # total control
                #u = u_ref + du # combine feedback and feedforward control twist components
                u = u_ref + du

                # update control gains for the next timestep
                self.k_n = 2*u[0]/(self.L**2) #kn
                self.k_g = u[0]/self.L #kg

                # ensure within performance limitation
                if u[0] > self.v_max: u[0] = self.v_max
                if u[0] < -self.v_max: u[0] = -self.v_max
                if u[1] > self.w_max: u[1] = self.w_max
                if u[1] < -self.w_max: u[1] = -self.w_max

                # actuator commands                 
                q = self.ddrive.inv_kinematics(u)            

                wheel_speed_msg = Vector3Stamped()
                wheel_speed_msg.vector.x = q[0,0] # Right wheelspeed rad/s
                wheel_speed_msg.vector.y = q[1,0] # Left wheelspeed rad/s

                self.cmd_wheelrate_right = wheel_speed_msg.vector.x
                self.cmd_wheelrate_left = wheel_speed_msg.vector.y
                
                ################################################################################

                # > Act < #
                # Send commands to the robot        
                self.wheel_speed_pub.publish(wheel_speed_msg)
                self.datalog.log(wheel_speed_msg, topic_name="/wheel_speeds_cmd")

            else:
                self.stop_to_think = False

            # Groundtruth
            if self.plotGroundtruth is not None:
                p_robot_truth = Vector(3)
                p_robot_truth[0,0] = self.groundtruth_northings
                p_robot_truth[1,0] = self.groundtruth_eastings
                p_robot_truth[2,0] = self.groundtruth_yaw
                self.p_groundtruth_tracker = p_robot_truth[0:3,0]

                dp_truth = self.state - p_robot_truth
                dp_truth[2] = (dp_truth[2] + np.pi) % (2 * np.pi) - np.pi # handle angle wrapping for yaw

                



if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument(
        "--time",
        type=float,
        default=-1,
        help="Time to run an experiment for. If negative, run forever.",
    )
    parser.add_argument(
        "--simulation",
        action="store_true",
        help="Run in simulation mode. Defaults to False",
    )

    args = parser.parse_args()

    laptop_pilot = LaptopPilot(args.simulation)
    laptop_pilot.run(args.time)