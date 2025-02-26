from laptop import LaptopPilot
import openpyxl
from openpyxl import load_workbook
from src.Libraries.math_feeg6043 import Vector, Transpose, m2l, change_to_list
from src.Libraries.model_feeg6043 import motion_model
import numpy as np
class createExcelFile:
    def __init__(self, filename = "reference.xlsx"):
        # Creates workbook and worksheet to modify
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.filename = filename
        self.workbook.save(filename)

        # Initialise variable to store data
        self.dataLine = []

    def extend_data(self, data):
        # adds to list self.dataLine
        data = change_to_list(data)
        self.dataLine.extend(data)
    def export_to_excel(self):
        # appends dataLine to sheet and saves file
        self.workbook = load_workbook(self.filename)
        self.worksheet.append(self.dataLine)       
        self.workbook.save(self.filename)
        self.dataLine = []
        self.workbook.close

def test_createExcelFile():
    testSheet = createExcelFile("MEGAGOONEAAAH.xlsx")
    p_example = Vector(3)
    p_example[0,0] = 1
    p_example[1,0] = 2
    p_example[2,0] = 3
    list2=[4,5,6]
    list3=[7,8,9]
    testSheet.extend_data([2,3])
    testSheet.extend_data([p_example[0,0]])
    testSheet.extend_data(list3)
    testSheet.export_to_excel()

def test_type_checker():
    array = np.array([[1, 2, 3], [4, 5, 6]])
    list = [1,2,3]
    matrix = Vector(4)
    assert type(array) == type(np.array([]))
    assert type(list) == type([])
    assert type(matrix) == type(np.array([]))
    
def test_motion_model():
    state = Vector(5)
    state[0] = 1
    state[1] = 2
    state[2] = 3
    state[3] = 0
    state[4] = 0

    u = Vector(2)
    u[0] = 2
    u[1] = 3

    dt = 2

    state, _, = motion_model(state,u,dt)